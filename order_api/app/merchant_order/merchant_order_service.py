import calendar
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import List, Optional, Tuple

from fastapi import HTTPException
from sqlalchemy.orm import Session, joinedload

from app.customer_order.customer_order_model import CustomerOrder
from app.customer_order.customer_order_service import (
    cancel_merchant_order,
    sync_customer_order_status,
)
from app.merchant.merchant_model import Merchant
from app.merchant_order.merchant_order_model import (
    MerchantOrder,
    MerchantOrderStatus,
    Notification,
    NotifikasiTipe,
    OrderItem,
)
from app.merchant_order.merchant_order_schema import (
    DashboardChartPoint,
    DashboardTransaction,
    MerchantDashboardSummary,
    MerchantOrderStatusUpdate,
    NotificationMarkRead,
)
from app.withdrawal.withdrawal_model import Withdrawal, WithdrawalStatus

# Nama hari ringkas (index = weekday(); Senin = 0).
_HARI = ["Sen", "Sel", "Rab", "Kam", "Jum", "Sab", "Min"]

# Nama bulan (index 1..12) untuk laporan tahunan.
_BULAN = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
          "Juli", "Agustus", "September", "Oktober", "November", "Desember"]

# Periode laporan penjualan yang didukung.
_PERIODE_LABEL = {"daily": "Harian", "weekly": "Mingguan", "monthly": "Bulanan", "yearly": "Tahunan"}

# Transisi status yang diizinkan
_ALLOWED_TRANSITIONS = {
    MerchantOrderStatus.BARU:       {MerchantOrderStatus.TERBUKA,  MerchantOrderStatus.DIBATALKAN},
    MerchantOrderStatus.TERBUKA:    {MerchantOrderStatus.DIPROSES, MerchantOrderStatus.DIBATALKAN},
    MerchantOrderStatus.DIPROSES:   {MerchantOrderStatus.SELESAI,  MerchantOrderStatus.DIBATALKAN},
    MerchantOrderStatus.SELESAI:    set(),
    MerchantOrderStatus.DIBATALKAN: set(),
}


# ── Eager load ────────────────────────────────────────────────────────────────

def _load_merchant_order(db: Session, order_id: int) -> MerchantOrder:
    """Ambil MerchantOrder beserta semua relasi yang dibutuhkan response."""
    order = (
        db.query(MerchantOrder)
        .options(
            joinedload(MerchantOrder.merchant),
            joinedload(MerchantOrder.customer_order)
                .joinedload(CustomerOrder.customer),
            joinedload(MerchantOrder.customer_order)
                .joinedload(CustomerOrder.dining_table),
            joinedload(MerchantOrder.items)
                .joinedload(OrderItem.product),
        )
        .filter(MerchantOrder.id == order_id)
        .first()
    )
    if not order:
        raise HTTPException(status_code=404, detail=f"Merchant order {order_id} tidak ditemukan")
    return order


# ── Query ─────────────────────────────────────────────────────────────────────

def get_merchant_order_or_404(db: Session, order_id: int) -> MerchantOrder:
    return _load_merchant_order(db, order_id)


def list_merchant_orders(
    db: Session,
    merchant_id: Optional[int] = None,
    status: Optional[MerchantOrderStatus] = None,
    offset: int = 0,
    limit: int = 50,
) -> List[MerchantOrder]:
    query = (
        db.query(MerchantOrder)
        .options(
            joinedload(MerchantOrder.merchant),
            joinedload(MerchantOrder.customer_order)
                .joinedload(CustomerOrder.customer),
            joinedload(MerchantOrder.customer_order)
                .joinedload(CustomerOrder.dining_table),
            joinedload(MerchantOrder.items)
                .joinedload(OrderItem.product),
        )
        .order_by(MerchantOrder.created_at.desc())
    )
    if merchant_id is not None:
        query = query.filter(MerchantOrder.merchant_id == merchant_id)
    if status:
        query = query.filter(MerchantOrder.status == status)
    return query.offset(offset).limit(limit).all()


# ── Update status ─────────────────────────────────────────────────────────────

def update_status(db: Session, order_id: int, data: MerchantOrderStatusUpdate) -> MerchantOrder:
    """Ubah status MerchantOrder lalu sinkronkan status CustomerOrder induk.

    Transisi yang diizinkan:
      baru       → terbuka | dibatalkan
      terbuka    → diproses | dibatalkan
      diproses   → selesai | dibatalkan
      selesai    → (final)
      dibatalkan → (final)

    Jika dibatalkan: stok dikembalikan otomatis.
    Setiap perubahan: notifikasi baru di inbox merchant.
    CustomerOrder diperbarui otomatis via sync_customer_order_status().
    """
    order = _load_merchant_order(db, order_id)
    new_status = data.status

    if new_status == order.status:
        return order

    if new_status not in _ALLOWED_TRANSITIONS[order.status]:
        raise HTTPException(
            status_code=400,
            detail=f"Transisi {order.status.value} → {new_status.value} tidak diizinkan",
        )

    # Pembatalan ditangani helper bersama: set status + kembalikan stok +
    # hitung ulang total struk + catat refund parsial bila sudah dibayar.
    if new_status == MerchantOrderStatus.DIBATALKAN:
        cancel_merchant_order(db, order)
    else:
        order.status = new_status

    # Buat notifikasi
    if new_status == MerchantOrderStatus.DIBATALKAN:
        tipe, judul = NotifikasiTipe.ORDER_DIBATALKAN, "Pesanan dibatalkan"
    elif new_status == MerchantOrderStatus.SELESAI:
        tipe, judul = NotifikasiTipe.ORDER_SELESAI, "Pesanan selesai"
    else:
        tipe, judul = NotifikasiTipe.STATUS_BERUBAH, "Status pesanan berubah"

    db.add(Notification(
        merchant_id=order.merchant_id,
        merchant_order_id=order.id,
        tipe=tipe,
        judul=judul,
        pesan=f"{order.order_code} → {new_status.value}",
    ))

    # Sinkronkan status CustomerOrder induk
    if order.customer_order:
        sync_customer_order_status(order.customer_order)

    db.commit()

    # Reload dengan joinedload agar semua relasi terisi di response
    return _load_merchant_order(db, order_id)


# ── Notifikasi ────────────────────────────────────────────────────────────────

def list_notifications(
    db: Session,
    merchant_id: int,
    only_unread: bool = False,
    offset: int = 0,
    limit: int = 50,
) -> List[Notification]:
    query = (
        db.query(Notification)
        .filter(Notification.merchant_id == merchant_id)
        .order_by(Notification.created_at.desc())
    )
    if only_unread:
        query = query.filter(Notification.is_read.is_(False))
    return query.offset(offset).limit(limit).all()


def mark_notifications_read(db: Session, merchant_id: int, data: NotificationMarkRead) -> dict:
    updated = (
        db.query(Notification)
        .filter(Notification.merchant_id == merchant_id, Notification.id.in_(data.ids))
        .update({Notification.is_read: True}, synchronize_session=False)
    )
    db.commit()
    return {"updated": updated}


def mark_all_notifications_read(db: Session, merchant_id: int) -> dict:
    updated = (
        db.query(Notification)
        .filter(Notification.merchant_id == merchant_id, Notification.is_read.is_(False))
        .update({Notification.is_read: True}, synchronize_session=False)
    )
    db.commit()
    return {"updated": updated}


# ── Dashboard Keuangan (Kontrol) ────────────────────────────────────────────────

def _local_date(dt: Optional[datetime]) -> Optional[date]:
    """Ambil komponen tanggal (lokal) dari datetime yang mungkin tz-aware."""
    if dt is None:
        return None
    if dt.tzinfo is not None:
        dt = dt.astimezone()
    return dt.date()


def get_dashboard_summary(db: Session, merchant: Merchant) -> MerchantDashboardSummary:
    """Ringkasan keuangan untuk halaman Kontrol merchant.

    Saldo & status toko diambil dari record merchant; angka pesanan/grafik/
    transaksi dihitung dari MerchantOrder + Withdrawal milik merchant ini.
    """
    today = date.today()

    orders = (
        db.query(MerchantOrder)
        .options(
            joinedload(MerchantOrder.items).joinedload(OrderItem.product),
            joinedload(MerchantOrder.customer_order).joinedload(CustomerOrder.customer),
        )
        .filter(MerchantOrder.merchant_id == merchant.id)
        .order_by(MerchantOrder.created_at.desc())
        .all()
    )

    PAID = {
        MerchantOrderStatus.TERBUKA,
        MerchantOrderStatus.DIPROSES,
        MerchantOrderStatus.SELESAI,
    }

    # ── Ringkasan hari ini ──────────────────────────────────────────────────
    total_order = 0
    total_pendapatan = 0.0
    produk_qty: dict[str, int] = {}

    # ── Grafik 7 hari terakhir (pendapatan pesanan selesai) ─────────────────
    week_days = [today - timedelta(days=i) for i in range(6, -1, -1)]
    week_revenue: dict[date, float] = {d: 0.0 for d in week_days}

    for o in orders:
        d = _local_date(o.created_at)
        if d == today and o.status in PAID:
            total_order += 1
        if o.status == MerchantOrderStatus.SELESAI:
            if d == today:
                total_pendapatan += o.total_harga or 0.0
                for it in o.items:
                    if it.product:
                        produk_qty[it.product.nama] = produk_qty.get(it.product.nama, 0) + it.jumlah
            if d in week_revenue:
                week_revenue[d] += o.total_harga or 0.0

    produk_terlaris = max(produk_qty, key=produk_qty.get) if produk_qty else None

    weekly_chart = [
        DashboardChartPoint(day=_HARI[d.weekday()], value=week_revenue[d])
        for d in week_days
    ]

    # ── Daftar transaksi: pesanan masuk + pencairan keluar ──────────────────
    transactions: List[DashboardTransaction] = []
    for o in orders:
        if o.status not in PAID:
            continue
        transactions.append(
            DashboardTransaction(
                id=o.order_code,
                name=o.pelanggan_nama or "Pelanggan",
                amount=o.total_harga or 0.0,
                type="masuk",
                created_at=o.created_at,
            )
        )

    withdrawals = (
        db.query(Withdrawal)
        .filter(
            Withdrawal.merchant_id == merchant.id,
            Withdrawal.status == WithdrawalStatus.APPROVED,
        )
        .all()
    )
    for w in withdrawals:
        transactions.append(
            DashboardTransaction(
                id=f"TRF-{w.id}",
                name="Pencairan Dana",
                amount=w.amount,
                type="keluar",
                created_at=w.processed_at or w.requested_at,
            )
        )

    transactions.sort(key=lambda t: t.created_at, reverse=True)
    transactions = transactions[:15]

    return MerchantDashboardSummary(
        saldo=merchant.balance,
        is_open=merchant.is_open,
        toko_nama=merchant.nama,
        lokasi=" · ".join(filter(None, [merchant.block, merchant.category])) or None,
        total_order=total_order,
        total_pendapatan=total_pendapatan,
        produk_terlaris=produk_terlaris,
        weekly_chart=weekly_chart,
        transactions=transactions,
    )


# ── Laporan Penjualan (export Excel) ─────────────────────────────────────────────

def _report_buckets(period: str, today: date) -> Tuple[List[Tuple[str, date, date]], date, date]:
    """Tentukan rentang & daftar bucket laporan sesuai periode.

    Return (buckets, periode_mulai, periode_selesai) dengan tiap bucket =
    (label, tanggal_mulai, tanggal_selesai) inklusif.
      - daily   : hari ini saja, satu bucket.
      - weekly  : 7 hari terakhir, bucket per hari.
      - monthly : bulan berjalan (tgl 1 s/d hari ini), bucket per hari.
      - yearly  : tahun berjalan (Jan s/d bulan ini), bucket per bulan.
    """
    if period == "daily":
        label = f"{_HARI[today.weekday()]}, {today.strftime('%d/%m/%Y')}"
        return [(label, today, today)], today, today

    if period == "weekly":
        start = today - timedelta(days=6)
        buckets = [
            (f"{_HARI[(start + timedelta(days=i)).weekday()]}, "
             f"{(start + timedelta(days=i)).strftime('%d/%m')}",
             start + timedelta(days=i), start + timedelta(days=i))
            for i in range(7)
        ]
        return buckets, start, today

    if period == "monthly":
        start = today.replace(day=1)
        days = (today - start).days + 1
        buckets = [
            (f"{(start + timedelta(days=i)).day} {_BULAN[start.month]}",
             start + timedelta(days=i), start + timedelta(days=i))
            for i in range(days)
        ]
        return buckets, start, today

    # yearly
    start = today.replace(month=1, day=1)
    buckets = []
    for m in range(1, today.month + 1):
        last_day = calendar.monthrange(today.year, m)[1]
        m_start = date(today.year, m, 1)
        m_end = date(today.year, m, last_day)
        buckets.append((f"{_BULAN[m]} {today.year}", m_start, m_end))
    return buckets, start, today


def build_sales_report(db: Session, merchant: Merchant, period: str) -> Tuple[bytes, str]:
    """Bangun file Excel (.xlsx) laporan penjualan merchant untuk satu periode.

    Hanya pesanan berstatus SELESAI yang dihitung sebagai penjualan (konsisten
    dengan perhitungan pendapatan di dashboard). Return (bytes_xlsx, filename).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if period not in _PERIODE_LABEL:
        raise HTTPException(status_code=400, detail="Periode harus daily, weekly, monthly, atau yearly")

    today = date.today()
    buckets, periode_mulai, periode_selesai = _report_buckets(period, today)

    orders = (
        db.query(MerchantOrder)
        .options(joinedload(MerchantOrder.items).joinedload(OrderItem.product))
        .filter(
            MerchantOrder.merchant_id == merchant.id,
            MerchantOrder.status == MerchantOrderStatus.SELESAI,
        )
        .all()
    )

    # Hanya pesanan dalam rentang periode.
    in_range = [
        o for o in orders
        if (d := _local_date(o.created_at)) is not None
        and periode_mulai <= d <= periode_selesai
    ]

    # Agregasi per bucket + ringkasan + produk terlaris.
    bucket_orders = {i: 0 for i in range(len(buckets))}
    bucket_revenue = {i: 0.0 for i in range(len(buckets))}
    produk_qty: dict[str, int] = {}
    total_order = 0
    total_pendapatan = 0.0

    for o in in_range:
        d = _local_date(o.created_at)
        for i, (_, b_start, b_end) in enumerate(buckets):
            if b_start <= d <= b_end:
                bucket_orders[i] += 1
                bucket_revenue[i] += o.total_harga or 0.0
                break
        total_order += 1
        total_pendapatan += o.total_harga or 0.0
        for it in o.items:
            if it.product:
                produk_qty[it.product.nama] = produk_qty.get(it.product.nama, 0) + it.jumlah

    produk_terlaris = max(produk_qty, key=produk_qty.get) if produk_qty else "-"
    label = _PERIODE_LABEL[period]

    # ── Susun workbook ────────────────────────────────────────────────────────
    BRAND = "1D3A27"
    GOLD = "C8961A"
    wb = Workbook()
    ws = wb.active
    ws.title = f"Laporan {label}"

    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color=BRAND)
    head_fill = PatternFill("solid", fgColor=BRAND)
    gold_fill = PatternFill("solid", fgColor=GOLD)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    rupiah_fmt = '"Rp"#,##0'
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22

    # Judul & meta
    ws["A1"] = "Laporan Penjualan"
    ws["A1"].font = title_font
    ws["A2"] = f"Periode: {label}"
    ws["A2"].font = bold
    ws["A3"] = f"Toko: {merchant.nama}"
    ws["A4"] = (f"Rentang: {periode_mulai.strftime('%d/%m/%Y')} "
                f"– {periode_selesai.strftime('%d/%m/%Y')}")
    ws["A5"] = f"Dicetak: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # Ringkasan
    ws["A7"] = "Ringkasan"
    ws["A7"].font = Font(bold=True, color=BRAND)
    summary = [
        ("Total Pesanan Selesai", total_order, None),
        ("Total Pendapatan", total_pendapatan, rupiah_fmt),
        ("Produk Terlaris", produk_terlaris, None),
    ]
    r = 8
    for nama_ringkas, val, fmt in summary:
        ws.cell(row=r, column=1, value=nama_ringkas).font = bold
        c = ws.cell(row=r, column=2, value=val)
        if fmt:
            c.number_format = fmt
        r += 1

    # Tabel rincian per bucket
    head_row = r + 1
    bucket_header = {"daily": "Tanggal", "weekly": "Tanggal", "monthly": "Tanggal", "yearly": "Bulan"}[period]
    headers = [bucket_header, "Jumlah Pesanan", "Pendapatan"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=head_row, column=col, value=h)
        cell.font = white_bold
        cell.fill = head_fill
        cell.border = border
        cell.alignment = center

    data_row = head_row + 1
    for i, (blabel, _, _) in enumerate(buckets):
        ws.cell(row=data_row, column=1, value=blabel).border = border
        oc = ws.cell(row=data_row, column=2, value=bucket_orders[i])
        oc.border = border
        oc.alignment = right
        rc = ws.cell(row=data_row, column=3, value=bucket_revenue[i])
        rc.number_format = rupiah_fmt
        rc.border = border
        rc.alignment = right
        data_row += 1

    # Baris total
    tc1 = ws.cell(row=data_row, column=1, value="TOTAL")
    tc1.font = bold
    tc1.fill = gold_fill
    tc1.border = border
    tc2 = ws.cell(row=data_row, column=2, value=total_order)
    tc2.font = bold
    tc2.fill = gold_fill
    tc2.border = border
    tc2.alignment = right
    tc3 = ws.cell(row=data_row, column=3, value=total_pendapatan)
    tc3.font = bold
    tc3.fill = gold_fill
    tc3.number_format = rupiah_fmt
    tc3.border = border
    tc3.alignment = right

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"laporan-penjualan-{period}-{today.strftime('%Y%m%d')}.xlsx"
    return buf.getvalue(), filename